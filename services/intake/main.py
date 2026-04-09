# =============================================================================
# POSEIDON Intake API — Port 8003
# EOB/835 parsing, CSV batch intake, OCR, document ingestion
# Data folder watcher for drop-in AI training files
# =============================================================================

from __future__ import annotations

import csv
import base64
import io
import json
import os
import hashlib
import re
import sys
import uuid
from datetime import datetime, timedelta, timezone
from email import message_from_bytes, policy
from email.header import decode_header
from pathlib import Path
from typing import Any

import httpx
import openpyxl
import pdfplumber
from fastapi import BackgroundTasks, Depends, File, Form, HTTPException, Request, UploadFile, status
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from psycopg.rows import dict_row  # type: ignore[import-untyped]

# Shared module: Docker has /app/shared; local uses repo/services/shared
_shared_dir = Path("/app/shared") if Path("/app/shared").exists() else (Path(__file__).resolve().parent.parent / "shared")
sys.path.insert(0, str(_shared_dir))
from base import create_app, get_redis, logger, settings

# ---------------------------------------------------------------------------
app = create_app(
    title="POSEIDON Intake API",
    version="2.0.0",
    description="EOB/835 ingestion, batch CSV intake, OCR parsing, data drop folder",
)

DATA_DIR = Path("/app/data")
EOB_DIR = DATA_DIR / "eobs"
DENIALS_DIR = DATA_DIR / "denials"
APPEALS_DIR = DATA_DIR / "appeals"
SPREADSHEETS_DIR = DATA_DIR / "spreadsheets"
TRAINING_DIR = DATA_DIR / "training"
PROCESSED_DIR = DATA_DIR / "processed"

for d in [EOB_DIR, DENIALS_DIR, APPEALS_DIR, SPREADSHEETS_DIR, TRAINING_DIR, PROCESSED_DIR]:
    d.mkdir(parents=True, exist_ok=True)


def sql(query: str) -> str:
    return re.sub(r"\$\d+", "%s", query)


async def fetch_one(conn, query: str, *params):
    async with conn.cursor(row_factory=dict_row) as cur:
        await cur.execute(sql(query), params)
        return await cur.fetchone()


async def fetch_all(conn, query: str, *params):
    async with conn.cursor(row_factory=dict_row) as cur:
        await cur.execute(sql(query), params)
        return await cur.fetchall()


async def exec_write(conn, query: str, *params) -> int:
    async with conn.cursor() as cur:
        await cur.execute(sql(query), params)
        return cur.rowcount


def _serialize(value: Any) -> Any:
    if isinstance(value, dict):
        return {k: _serialize(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_serialize(v) for v in value]
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            return str(value)
    return value


async def _record_workflow_event(conn, org_id: str, event_type: str, payload: dict[str, Any], order_id: str | None = None):
    await exec_write(
        conn,
        """
        INSERT INTO workflow_events (org_id, order_id, event_type, payload)
        VALUES ($1,$2,$3,$4)
        """,
        org_id,
        order_id,
        event_type,
        json.dumps(payload),
    )


def _json_hash(payload: dict[str, Any]) -> str:
    encoded = json.dumps(payload, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def _request_idempotency_key(request: Request) -> str | None:
    value = request.headers.get("Idempotency-Key")
    if not value:
        return None
    normalized = value.strip()
    return normalized or None


# ---------------------------------------------------------------------------
# EOB/835 Parser
# ---------------------------------------------------------------------------

class EOBParser:
    """Parse CMS EOB PDFs and 835 EDI transaction sets."""

    # CARC → denial category mapping
    CARC_CATEGORIES = {
        # Eligibility
        "1": "eligibility", "2": "eligibility", "27": "eligibility",
        "31": "eligibility", "96": "eligibility",
        # Medical necessity
        "50": "medical_necessity", "57": "medical_necessity",
        "151": "medical_necessity", "167": "medical_necessity",
        # Authorization
        "4": "authorization", "15": "authorization", "197": "authorization",
        # Coding
        "6": "coding", "16": "coding", "29": "coding",
        # Timely filing
        "29": "timely_filing", "119": "timely_filing",
        # Duplicate
        "18": "duplicate", "97": "duplicate",
        # COB
        "22": "coordination", "23": "coordination",
    }

    def parse_835_edi(self, content: str) -> dict:
        """Parse X12 835 EDI transaction set."""
        segments = content.strip().split("~")
        claims = []
        current_claim: dict = {}
        adjustments: list = []

        for seg in segments:
            elements = seg.strip().split("*")
            seg_id = elements[0] if elements else ""

            if seg_id == "CLP":
                if current_claim:
                    current_claim["adjustments"] = adjustments
                    claims.append(current_claim)
                current_claim = {
                    "claim_id": elements[1] if len(elements) > 1 else "",
                    "claim_status": elements[2] if len(elements) > 2 else "",
                    "billed_amount": float(elements[3]) if len(elements) > 3 else 0,
                    "paid_amount": float(elements[4]) if len(elements) > 4 else 0,
                    "patient_responsibility": float(elements[5]) if len(elements) > 5 else 0,
                    "payer_claim_number": elements[7] if len(elements) > 7 else "",
                }
                adjustments = []

            elif seg_id == "CAS":
                # Claim adjustment
                adj_group = elements[1] if len(elements) > 1 else ""
                i = 2
                while i + 1 < len(elements):
                    carc = elements[i] if i < len(elements) else ""
                    amount = float(elements[i + 1]) if i + 1 < len(elements) else 0
                    if carc:
                        adjustments.append({
                            "group": adj_group,
                            "carc": carc,
                            "amount": amount,
                            "category": self.CARC_CATEGORIES.get(carc, "other"),
                        })
                    i += 3  # CARC, amount, quantity triplets

            elif seg_id == "NM1" and len(elements) > 3:
                if elements[1] == "QC":  # Patient
                    current_claim["patient_last"] = elements[3]
                    current_claim["patient_first"] = elements[4] if len(elements) > 4 else ""
                elif elements[1] == "82":  # Rendering provider
                    current_claim["provider_npi"] = elements[9] if len(elements) > 9 else ""

            elif seg_id == "DTM" and len(elements) > 2:
                if elements[1] == "405":  # Service date
                    current_claim["service_date"] = elements[2]

            elif seg_id == "SVC" and len(elements) > 1:
                # Service line
                hcpcs = elements[1].split(":")[1] if ":" in elements[1] else elements[1]
                current_claim.setdefault("service_lines", []).append({
                    "hcpcs": hcpcs,
                    "billed": float(elements[2]) if len(elements) > 2 else 0,
                    "paid": float(elements[3]) if len(elements) > 3 else 0,
                })

        if current_claim:
            current_claim["adjustments"] = adjustments
            claims.append(current_claim)

        denials = [c for c in claims if float(c.get("paid_amount", 0)) == 0 and c.get("adjustments")]

        return {
            "total_claims": len(claims),
            "total_paid": sum(c.get("paid_amount", 0) for c in claims),
            "total_denied": len(denials),
            "total_billed": sum(c.get("billed_amount", 0) for c in claims),
            "claims": claims,
            "denials": denials,
            "parsed_at": datetime.now(timezone.utc).isoformat(),
        }

    def parse_eob_pdf(self, pdf_bytes: bytes) -> dict:
        """Extract key fields from PDF EOB."""
        result: dict[str, Any] = {
            "claims": [],
            "raw_text_snippet": "",
            "parsed_at": datetime.now(timezone.utc).isoformat(),
        }
        try:
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)
                result["raw_text_snippet"] = full_text[:500]

                # Extract claim numbers
                claim_nums = re.findall(r"Claim\s+(?:Number|#)[:\s]+([A-Z0-9\-]+)", full_text, re.I)
                # Extract amounts
                amounts = re.findall(r"\$\s*([\d,]+\.\d{2})", full_text)
                # Extract CARC codes
                carcs = re.findall(r"\bCARC[:\s]+(\d+)\b", full_text, re.I)
                # Extract RARC codes
                rarcs = re.findall(r"\bRARC[:\s]+([A-Z]\d+)\b", full_text, re.I)

                result["claim_numbers"] = claim_nums
                result["amounts_found"] = [float(a.replace(",", "")) for a in amounts[:10]]
                result["carc_codes"] = list(set(carcs))
                result["rarc_codes"] = list(set(rarcs))
                result["denial_categories"] = list({
                    self.CARC_CATEGORIES.get(c, "other") for c in carcs
                })
        except Exception as e:
            logger.warning("EOB PDF parse error: %s", e)
            result["parse_error"] = str(e)
        return result


eob_parser = EOBParser()


# ---------------------------------------------------------------------------
# CSV/Excel Batch Intake
# ---------------------------------------------------------------------------

class BatchIntakeProcessor:
    REQUIRED_COLUMNS = {"patient_id", "hcpcs_code", "payer_id", "diagnosis_code"}
    OPTIONAL_COLUMNS = {"first_name", "last_name", "dob", "insurance_id", "auth_number", "npi"}

    def process_csv(self, content: str) -> dict:
        reader = csv.DictReader(io.StringIO(content))
        headers = set(reader.fieldnames or [])
        missing = self.REQUIRED_COLUMNS - headers
        if missing:
            return {
                "error": f"Missing required columns: {', '.join(missing)}",
                "required": list(self.REQUIRED_COLUMNS),
                "found": list(headers),
            }

        rows = []
        errors = []
        for i, row in enumerate(reader, 1):
            cleaned, errs = self._validate_row(row, i)
            if errs:
                errors.extend(errs)
            else:
                rows.append(cleaned)

        return {
            "total_rows": i if rows or errors else 0,
            "valid_rows": len(rows),
            "error_rows": len(errors),
            "errors": errors[:20],  # cap errors in response
            "records": rows,
        }

    def _validate_row(self, row: dict, line_num: int) -> tuple[dict, list]:
        errors = []
        cleaned = {}

        # HCPCS validation
        hcpcs = row.get("hcpcs_code", "").strip().upper()
        if not re.match(r"^[A-Z]\d{4}$", hcpcs):
            errors.append({"line": line_num, "field": "hcpcs_code", "value": hcpcs, "error": "Invalid HCPCS format"})
        else:
            cleaned["hcpcs_code"] = hcpcs

        # ICD-10 validation
        dx = row.get("diagnosis_code", "").strip().upper().replace(".", "")
        if not (3 <= len(dx) <= 7):
            errors.append({"line": line_num, "field": "diagnosis_code", "value": dx, "error": "Invalid ICD-10 format"})
        else:
            cleaned["diagnosis_code"] = dx

        cleaned["patient_id"] = row.get("patient_id", "").strip()
        cleaned["payer_id"] = row.get("payer_id", "").strip().upper()
        cleaned["line"] = line_num

        # Optional fields
        for col in self.OPTIONAL_COLUMNS:
            if col in row and row[col]:
                cleaned[col] = row[col].strip()

        return cleaned, errors


batch_processor = BatchIntakeProcessor()


# ---------------------------------------------------------------------------
# Email Intake Workflow
# ---------------------------------------------------------------------------

class EmailPollRequest(BaseModel):
    max_messages: int = 10
    mailbox: str = "INBOX"
    mark_seen: bool = True
    search: str = "UNSEEN"


class LegacyPatientIntakePayload(BaseModel):
    source: str
    payload: dict[str, Any]
    submitted_at: str | None = None
    idempotency_key: str | None = None


def _decode_mime_header(value: str | None) -> str:
    if not value:
        return ""
    parts: list[str] = []
    for chunk, encoding in decode_header(value):
        if isinstance(chunk, bytes):
            parts.append(chunk.decode(encoding or "utf-8", errors="replace"))
        else:
            parts.append(chunk)
    return "".join(parts).strip()


def _normalize_key(value: str) -> str:
    return value.lower().replace("/", "_").replace("-", "_").replace(" ", "_")


def _normalize_date(value: str | None) -> str:
    raw = (value or "").strip()
    if not raw:
        return "1970-01-01"
    for fmt in (
        "%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y", "%m-%d-%y",
        "%b %d, %Y", "%b %d %Y", "%B %d, %Y", "%B %d %Y",
    ):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(raw).date().isoformat()
    except ValueError:
        return "1970-01-01"


def _split_codes(value: str) -> list[str]:
    return [item.strip().upper().replace(".", "") for item in re.split(r"[,;\s]+", value) if item.strip()]


ICD10_TO_HCPCS_CONTEXT: tuple[tuple[tuple[str, ...], list[str], str], ...] = (
    (("M17", "M23", "M2556", "M2557"), ["L1833"], "knee-bracing"),
    (("M16", "M2445", "M2555"), ["L1686"], "hip-bracing"),
    (("G82", "G80", "G35", "R26"), ["K0823"], "mobility"),
    (("M54", "M51", "M48"), ["L0650"], "lumbar-bracing"),
)


def _infer_hcpcs_from_diagnosis(diagnosis_codes: list[str]) -> tuple[list[str], str]:
    normalized = [str(code or "").upper().replace(".", "") for code in diagnosis_codes if str(code or "").strip()]
    for prefixes, hcpcs_codes, source in ICD10_TO_HCPCS_CONTEXT:
        for code in normalized:
            if any(code.startswith(prefix) for prefix in prefixes):
                return list(hcpcs_codes), source
    return [], "no-inference"


def _split_name(full_name: str) -> tuple[str, str]:
    cleaned = full_name.strip()
    if not cleaned:
        return "Unknown", "Patient"
    if "," in cleaned:
        last, first = [part.strip() for part in cleaned.split(",", 1)]
        return first or "Unknown", last or "Patient"
    parts = cleaned.split()
    if len(parts) == 1:
        return parts[0], "Patient"
    return parts[0], " ".join(parts[1:])


def _extract_first(text: str, patterns: list[str]) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, re.I | re.M)
        if match:
            return match.group(1).strip()
    return ""


def _extract_pdf_text(pdf_bytes: bytes) -> str:
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        return "\n".join(page.extract_text() or "" for page in pdf.pages)


def _parse_pdf_attachment(filename: str, content: bytes, message_meta: dict[str, Any]) -> list[dict[str, Any]]:
    text = _extract_pdf_text(content)
    patient_name = _extract_first(
        text,
        [
            r"Patient(?:\s+Name)?[:\s]+([A-Z][A-Z,\-'\s]+)",
            r"Member(?:\s+Name)?[:\s]+([A-Z][A-Z,\-'\s]+)",
            r"Name[:\s]+([A-Z][A-Z,\-'\s]+)",
        ],
    )
    first_name, last_name = _split_name(patient_name)
    dob = _normalize_date(
        _extract_first(text, [r"(?:DOB|Date of Birth)[:\s]+([0-9/\-]{6,10})"])
    )
    insurance_id = _extract_first(
        text,
        [
            r"(?:Member\s*ID|Subscriber\s*ID|Insurance\s*ID)[:\s]+([A-Z0-9\-]+)",
            r"ID[:\s]+([A-Z0-9\-]{5,})",
        ],
    )
    email = _extract_first(
        text,
        [r"(?:Email|E-mail)[:\s]+([A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,})"],
    )
    payer = _extract_first(
        text,
        [
            r"(?:Insurance|Payer|Plan)[:\s]+([A-Z][A-Z0-9&.,'\/\-\s]+)",
        ],
    )
    hcpcs_codes = _split_codes(" ".join(re.findall(r"\b[A-Z]\d{4}\b", text)))
    diagnosis_codes = _split_codes(" ".join(re.findall(r"\b[A-Z]\d{2}(?:\.\d{1,4})?\b", text)))
    priority = "urgent" if re.search(r"\b(?:urgent|stat|expedite)\b", text, re.I) else "standard"

    inferred_note = None
    if not hcpcs_codes:
        hcpcs_codes, inferred_from = _infer_hcpcs_from_diagnosis(diagnosis_codes)
        inferred_note = f"hcpcs_inferred_from_icd10:{inferred_from}"
        logger.warning("Email PDF intake missing HCPCS; inferred %s from %s", ",".join(hcpcs_codes), inferred_from)

    payload = {
        "patient_name": patient_name or f"{first_name} {last_name}".strip(),
        "first_name": first_name,
        "last_name": last_name,
        "dob": dob,
        "email": email.lower() if email else None,
        "insurance_id": insurance_id,
        "payer": payer,
        "hcpcs_codes": hcpcs_codes,
        "diagnosis_codes": diagnosis_codes,
        "priority": priority,
        "notes": f"Email intake PDF: {filename}",
        "source_channel": "email_pdf",
        "source_reference": message_meta.get("message_id") or filename,
        "intake_payload": {
            "email_subject": message_meta.get("subject"),
            "email_from": message_meta.get("from"),
            "attachment_name": filename,
            "text_preview": text[:1200],
            "warnings": [inferred_note] if inferred_note else [],
        },
    }

    if not payload["patient_name"].strip():
        return []
    return [payload]


def _parse_csv_attachment(filename: str, content: bytes, message_meta: dict[str, Any]) -> list[dict[str, Any]]:
    text = content.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    orders: list[dict[str, Any]] = []
    for raw_row in reader:
        row = {_normalize_key(key): (value or "").strip() for key, value in raw_row.items()}
        patient_name = row.get("patient_name") or row.get("name") or " ".join(
            part for part in [row.get("first_name", ""), row.get("last_name", "")] if part
        ).strip()
        if not patient_name:
            continue
        first_name, last_name = _split_name(patient_name)
        hcpcs_raw = row.get("hcpcs") or row.get("hcpcs_code") or row.get("hcpcs_codes") or ""
        dx_raw = row.get("icd") or row.get("diagnosis_code") or row.get("diagnosis_codes") or ""
        diagnosis_codes = _split_codes(dx_raw)
        hcpcs_codes = _split_codes(hcpcs_raw)
        inferred_note = None
        if not hcpcs_codes:
            hcpcs_codes, inferred_from = _infer_hcpcs_from_diagnosis(diagnosis_codes)
            inferred_note = f"hcpcs_inferred_from_icd10:{inferred_from}"
            logger.warning("Email CSV intake missing HCPCS; inferred %s from %s", ",".join(hcpcs_codes), inferred_from)

        orders.append({
            "patient_name": patient_name,
            "first_name": row.get("first_name") or first_name,
            "last_name": row.get("last_name") or last_name,
            "dob": _normalize_date(row.get("dob") or row.get("date_of_birth")),
            "email": (row.get("email") or row.get("patient_email") or "").lower() or None,
            "insurance_id": row.get("insurance_id") or row.get("member_id") or row.get("subscriber_id"),
            "payer": row.get("payer") or row.get("payer_name") or row.get("insurance"),
            "payer_id": row.get("payer_id"),
            "hcpcs": hcpcs_raw,
            "hcpcs_codes": hcpcs_codes,
            "icd": dx_raw,
            "diagnosis_codes": diagnosis_codes,
            "npi": row.get("npi") or row.get("provider_npi"),
            "referring_physician_npi": row.get("referring_physician_npi"),
            "priority": row.get("priority") or "standard",
            "assigned_to_email": row.get("assigned_to_email") or row.get("owner_email"),
            "notes": row.get("notes") or f"Email intake spreadsheet: {filename}",
            "source_channel": "email_spreadsheet",
            "source_reference": message_meta.get("message_id") or filename,
            "intake_payload": {
                "email_subject": message_meta.get("subject"),
                "email_from": message_meta.get("from"),
                "attachment_name": filename,
                "raw_row": raw_row,
                "warnings": [inferred_note] if inferred_note else [],
            },
        })
    return orders


def _parse_xlsx_attachment(filename: str, content: bytes, message_meta: dict[str, Any]) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_normalize_key(str(value or "")) for value in rows[0]]
    csv_like = io.StringIO()
    writer = csv.DictWriter(csv_like, fieldnames=headers)
    writer.writeheader()
    for values in rows[1:]:
        writer.writerow({headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))})
    return _parse_csv_attachment(filename, csv_like.getvalue().encode("utf-8"), message_meta)


def _parse_attachment(filename: str, content: bytes, message_meta: dict[str, Any]) -> list[dict[str, Any]]:
    lower = filename.lower()
    if lower.endswith(".pdf"):
        return _parse_pdf_attachment(filename, content, message_meta)
    if lower.endswith(".csv"):
        return _parse_csv_attachment(filename, content, message_meta)
    if lower.endswith(".xlsx"):
        return _parse_xlsx_attachment(filename, content, message_meta)
    return []


async def _core_access_token() -> tuple[str, str]:
    core_url = settings.core_url.rstrip("/")
    email = os.getenv("CORE_API_EMAIL", "").strip()
    password = os.getenv("CORE_API_PASSWORD", "").strip()
    if not email or not password:
        raise HTTPException(status_code=500, detail="Missing CORE_API_EMAIL or CORE_API_PASSWORD.")

    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.post(
            f"{core_url}/auth/login",
            json={"email": email, "password": password},
        )
    if response.status_code != 200:
        raise HTTPException(status_code=502, detail=f"Core auth failed: {response.status_code}")
    payload = response.json()
    token = payload.get("access_token")
    if not token:
        raise HTTPException(status_code=502, detail="Core auth returned no access token.")
    return core_url, token


def _core_internal_headers(token: str) -> dict[str, str]:
    internal_api_key = settings.internal_api_key.strip()
    if not internal_api_key:
        raise HTTPException(status_code=500, detail="Missing INTERNAL_API_KEY for Core service requests.")
    return {
        "Authorization": f"Bearer {token}",
        "X-Internal-API-Key": internal_api_key,
    }


async def _submit_orders_to_core(orders: list[dict[str, Any]]) -> dict[str, Any]:
    core_url, token = await _core_access_token()
    async with httpx.AsyncClient(timeout=60.0) as client:
        response = await client.post(
            f"{core_url}/orders/import",
            json={"orders": orders},
            headers=_core_internal_headers(token),
        )
    if response.status_code not in {200, 201}:
        detail = response.text[:500]
        raise HTTPException(status_code=502, detail=f"Core import failed: {response.status_code} {detail}")
    return response.json()


async def _advance_core_order_workflow(order_id: str) -> dict[str, Any]:
    core_url, token = await _core_access_token()
    async with httpx.AsyncClient(timeout=60.0) as client:
        response = await client.post(
            f"{core_url}/workflow/orders/{order_id}/advance-from-intake",
            json={"auto_request_swo": True},
            headers=_core_internal_headers(token),
        )
    if response.status_code not in {200, 201}:
        detail = response.text[:500]
        raise HTTPException(status_code=502, detail=f"Core workflow advance failed: {response.status_code} {detail}")
    return response.json()


def _iter_email_attachments(message) -> list[tuple[str, bytes]]:
    attachments: list[tuple[str, bytes]] = []
    for part in message.walk():
        filename = part.get_filename()
        if not filename:
            continue
        decoded_name = _decode_mime_header(filename)
        payload = part.get_payload(decode=True)
        if payload:
            attachments.append((decoded_name, payload))
    return attachments


async def _gmail_access_token() -> tuple[str, str]:
    client_id = os.getenv("GMAIL_OAUTH_CLIENT_ID", "").strip()
    client_secret = os.getenv("GMAIL_OAUTH_CLIENT_SECRET", "").strip()
    refresh_token = os.getenv("GMAIL_OAUTH_REFRESH_TOKEN", "").strip()
    gmail_user = os.getenv("GMAIL_INTAKE_USER", "").strip() or os.getenv("EMAIL_INTAKE_USERNAME", "").strip()
    token_url = os.getenv("GMAIL_TOKEN_URL", "https://oauth2.googleapis.com/token").strip()

    if not client_id or not client_secret or not refresh_token or not gmail_user:
        raise HTTPException(
            status_code=500,
            detail="Missing Gmail OAuth configuration. Expected GMAIL_OAUTH_CLIENT_ID, GMAIL_OAUTH_CLIENT_SECRET, GMAIL_OAUTH_REFRESH_TOKEN, and GMAIL_INTAKE_USER.",
        )

    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.post(
            token_url,
            data={
                "client_id": client_id,
                "client_secret": client_secret,
                "refresh_token": refresh_token,
                "grant_type": "refresh_token",
            },
            headers={"Accept": "application/json"},
        )

    if response.status_code >= 400:
        raise HTTPException(status_code=502, detail=f"Gmail OAuth token request failed: {response.status_code}")

    payload = response.json()
    access_token = payload.get("access_token")
    if not access_token:
        raise HTTPException(status_code=502, detail="Gmail OAuth token response missing access_token.")

    return gmail_user, access_token


def _gmail_query(search: str) -> str:
    normalized = (search or "").strip()
    if not normalized or normalized.upper() == "UNSEEN":
        return "is:unread"
    return normalized


async def _fetch_email_messages(mailbox: str, search: str, max_messages: int) -> list[dict[str, Any]]:
    gmail_user, access_token = await _gmail_access_token()
    query = _gmail_query(search)
    list_url = f"https://gmail.googleapis.com/gmail/v1/users/{gmail_user}/messages"

    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.get(
            list_url,
            params={"q": query, "maxResults": max_messages},
            headers={"Authorization": f"Bearer {access_token}"},
        )

        if response.status_code >= 400:
            raise HTTPException(status_code=502, detail=f"Gmail list messages failed: {response.status_code}")

        payload = response.json()
        message_refs = payload.get("messages") or []
        messages: list[dict[str, Any]] = []

        for item in message_refs:
            gmail_id = str(item.get("id") or "")
            if not gmail_id:
                continue

            message_response = await client.get(
                f"{list_url}/{gmail_id}",
                params={"format": "raw"},
                headers={"Authorization": f"Bearer {access_token}"},
            )
            if message_response.status_code >= 400:
                continue

            message_payload = message_response.json()
            raw_value = message_payload.get("raw")
            if not raw_value:
                continue

            raw_bytes = base64.urlsafe_b64decode(raw_value + "=" * (-len(raw_value) % 4))
            message = message_from_bytes(raw_bytes, policy=policy.default)
            messages.append({
                "imap_id": gmail_id,
                "gmail_id": gmail_id,
                "mailbox": mailbox,
                "message_id": _decode_mime_header(message.get("Message-ID")),
                "subject": _decode_mime_header(message.get("Subject")),
                "from": _decode_mime_header(message.get("From")),
                "date": _decode_mime_header(message.get("Date")),
                "attachments": _iter_email_attachments(message),
            })

        return messages


async def _mark_email_messages_seen(mailbox: str, message_ids: list[str]) -> None:
    if not message_ids:
        return

    gmail_user, access_token = await _gmail_access_token()
    async with httpx.AsyncClient(timeout=30.0) as client:
        for message_id in message_ids:
            await client.post(
                f"https://gmail.googleapis.com/gmail/v1/users/{gmail_user}/messages/{message_id}/modify",
                json={"removeLabelIds": ["UNREAD"]},
                headers={"Authorization": f"Bearer {access_token}"},
            )


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.post("/ingest/eob")
async def ingest_eob(
    request: Request,
    file: UploadFile = File(...),
):
    """Accept EOB PDF or 835 EDI file. Auto-detect format."""
    content = await file.read()
    filename = file.filename or "unknown"

    # Save to data/eobs/
    save_path = EOB_DIR / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
    save_path.write_bytes(content)

    if filename.lower().endswith(".pdf"):
        result = eob_parser.parse_eob_pdf(content)
        result["source_file"] = filename
        result["format"] = "pdf"
    elif filename.lower().endswith((".835", ".txt", ".edi")):
        text = content.decode("utf-8", errors="replace")
        result = eob_parser.parse_835_edi(text)
        result["source_file"] = filename
        result["format"] = "x12_835"
    else:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file type. Accepted: .pdf, .835, .txt, .edi",
        )

    # Push to ML training queue
    redis = get_redis(request)
    await redis.rpush("intake:eob_processed", json.dumps({
        "file": filename,
        "total_claims": result.get("total_claims", 0),
        "total_denied": result.get("total_denied", 0),
    }))

    logger.info("EOB ingested: %s format=%s", filename, result.get("format"))
    return result


@app.post("/ingest/batch")
async def ingest_batch(
    request: Request,
    file: UploadFile = File(...),
):
    """Batch CSV/Excel intake for patient/order creation."""
    content = await file.read()
    filename = file.filename or "batch.csv"

    save_path = SPREADSHEETS_DIR / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
    save_path.write_bytes(content)

    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8", errors="replace")
        result = batch_processor.process_csv(text)
    else:
        raise HTTPException(status_code=400, detail="Accepted formats: .csv")

    result["source_file"] = filename
    result["saved_to"] = str(save_path)
    return result


@app.post("/ingest/denial-file")
async def ingest_denial_file(
    request: Request,
    file: UploadFile = File(...),
):
    """Accept denial spreadsheets/CSVs for worklist population and ML training."""
    content = await file.read()
    filename = file.filename or "denials.csv"
    save_path = DENIALS_DIR / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
    save_path.write_bytes(content)

    rows_processed = 0
    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows_processed = sum(1 for _ in reader)

    redis = get_redis(request)
    await redis.rpush("intake:denial_file_uploaded", json.dumps({
        "file": filename,
        "path": str(save_path),
        "rows": rows_processed,
    }))

    return {
        "status": "ingested",
        "file": filename,
        "rows_detected": rows_processed,
        "saved_to": str(save_path),
        "queued_for_ml": True,
    }


@app.post("/ingest/appeal")
async def ingest_appeal(
    request: Request,
    file: UploadFile = File(...),
):
    """Accept appeal documents for tracking and pattern analysis."""
    content = await file.read()
    filename = file.filename or "appeal.pdf"
    save_path = APPEALS_DIR / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
    save_path.write_bytes(content)

    return {
        "status": "ingested",
        "file": filename,
        "saved_to": str(save_path),
        "next_steps": ["review_in_dashboard", "link_to_order"],
    }


@app.post("/ingest/email-intake/poll")
async def poll_email_intake(request: Request, payload: EmailPollRequest):
    """Poll the intake mailbox, download attachments, create patients/orders in Core, and assign work."""
    messages = await _fetch_email_messages(payload.mailbox, payload.search, payload.max_messages)

    processed_messages = []
    total_orders = 0
    total_attachments = 0

    successfully_processed_ids: list[str] = []
    failed_messages = 0

    for message in messages:
        message_orders: list[dict[str, Any]] = []
        processed_attachments: list[dict[str, Any]] = []
        message_status = "processed"
        message_failure: str | None = None

        try:
            for filename, content in message["attachments"]:
                total_attachments += 1
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                save_path = PROCESSED_DIR / f"{timestamp}_{filename}"
                save_path.write_bytes(content)

                parsed_orders = _parse_attachment(filename, content, message)
                if parsed_orders:
                    message_orders.extend(parsed_orders)
                    processed_attachments.append({
                        "filename": filename,
                        "saved_to": str(save_path),
                        "orders_detected": len(parsed_orders),
                    })

            import_result = None
            workflow_results: list[dict[str, Any]] = []
            if message_orders:
                import_result = await _submit_orders_to_core(message_orders)
                total_orders += len(message_orders)
                for item in import_result.get("results", []):
                    if item.get("status") == "created" and item.get("order_id"):
                        try:
                            workflow_results.append(
                                await _advance_core_order_workflow(str(item["order_id"]))
                            )
                        except HTTPException as exc:
                            workflow_results.append({
                                "order_id": item["order_id"],
                                "status": "workflow_error",
                                "detail": exc.detail,
                            })
            successfully_processed_ids.append(message["imap_id"])
        except Exception as exc:
            failed_messages += 1
            message_status = "failed"
            message_failure = str(exc)[:300]
            import_result = None
            workflow_results = []

        processed_messages.append({
            "imap_id": message["imap_id"],
            "message_id": message["message_id"],
            "subject": message["subject"],
            "from": message["from"],
            "attachments_processed": processed_attachments,
            "orders_submitted": len(message_orders),
            "status": message_status,
            "failure_reason": message_failure,
            "import_result": import_result,
            "workflow_results": workflow_results,
        })

    redis = get_redis(request)
    await redis.rpush("intake:email_poll_completed", json.dumps({
        "messages": len(processed_messages),
        "attachments": total_attachments,
        "orders_submitted": total_orders,
        "failed_messages": failed_messages,
        "mailbox": payload.mailbox,
    }))

    if payload.mark_seen:
        await _mark_email_messages_seen(
            payload.mailbox,
            successfully_processed_ids,
        )

    return {
        "status": "processed",
        "mailbox": payload.mailbox,
        "messages_checked": len(messages),
        "attachments_processed": total_attachments,
        "orders_submitted": total_orders,
        "failed_messages": failed_messages,
        "messages": processed_messages,
    }


@app.get("/data/inventory")
async def data_inventory():
    """List all files in the data drop folders."""
    def scan_dir(d: Path) -> dict:
        files = list(d.glob("*")) if d.exists() else []
        return {
            "count": len(files),
            "files": [
                {
                    "name": f.name,
                    "size_kb": round(f.stat().st_size / 1024, 1),
                    "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
                }
                for f in sorted(files, key=lambda x: x.stat().st_mtime, reverse=True)[:20]
            ],
        }

    return {
        "eobs": scan_dir(EOB_DIR),
        "denials": scan_dir(DENIALS_DIR),
        "appeals": scan_dir(APPEALS_DIR),
        "spreadsheets": scan_dir(SPREADSHEETS_DIR),
        "training": scan_dir(TRAINING_DIR),
        "processed": scan_dir(PROCESSED_DIR),
    }


@app.post("/patient-intake")
async def patient_intake(request: Request, payload: LegacyPatientIntakePayload):
    """Mommy Care Kit and other intake form submissions."""
    if not payload.source.strip():
        raise HTTPException(status_code=400, detail="source is required")
    if not payload.payload:
        raise HTTPException(status_code=400, detail="payload is required")

    normalized = payload.model_dump()
    idempotency_key = payload.idempotency_key or _request_idempotency_key(request) or _json_hash(normalized)
    intake_id = str(uuid.uuid4())
    received_at = datetime.now(timezone.utc).isoformat()

    redis = get_redis(request)
    dedup_key = f"intake:patient_form:{idempotency_key}"
    accepted = await redis.set(dedup_key, intake_id, ex=24 * 60 * 60, nx=True)
    if not accepted:
        prior = await redis.get(dedup_key)
        return {
            "intake_id": prior or intake_id,
            "status": "duplicate_ignored",
            "idempotent_replay": True,
            "received_at": received_at,
        }

    # Save raw intake
    save_path = DATA_DIR / "training" / f"intake_{intake_id}.json"
    save_path.write_text(
        json.dumps(
            {
                "intake_id": intake_id,
                "source": payload.source,
                "submitted_at": payload.submitted_at,
                "received_at": received_at,
                "idempotency_key": idempotency_key,
                "processing_status": "received",
                "failure_reason": None,
                "payload": payload.payload,
            },
            indent=2,
        )
    )

    await redis.rpush("intake:patient_form", json.dumps({
        "intake_id": intake_id,
        "source": payload.source,
        "received_at": received_at,
        "idempotency_key": idempotency_key,
        "payload_keys": list(payload.payload.keys()),
    }))

    return {"intake_id": intake_id, "status": "received", "idempotent_replay": False, "received_at": received_at}


# ---------------------------------------------------------------------------
# Canonical /api/v1 Intake compatibility layer
# ---------------------------------------------------------------------------

class IntakeInsurancePayload(BaseModel):
    payer_name: str
    payer_id: str | None = None
    member_id: str | None = None
    group_number: str | None = None
    subscriber_name: str | None = None
    subscriber_dob: str | None = None
    relationship: str = "self"
    is_primary: bool = True


class IntakePhysicianPayload(BaseModel):
    npi: str | None = None
    first_name: str | None = None
    last_name: str | None = None
    specialty: str | None = None
    phone: str | None = None
    fax: str | None = None
    facility_name: str | None = None


class IntakeDiagnosisPayload(BaseModel):
    icd10_code: str
    description: str | None = None
    is_primary: bool = False
    sequence: int = 1


class IntakeLineItemPayload(BaseModel):
    hcpcs_code: str
    modifier: str | None = None
    description: str | None = None
    quantity: int = 1
    unit_price: float | None = None
    billed_amount: float | None = None


class IntakePatientPayload(BaseModel):
    org_id: str
    first_name: str
    last_name: str
    date_of_birth: str
    gender: str | None = None
    phone: str | None = None
    email: str | None = None
    address_line1: str | None = None
    address_line2: str | None = None
    city: str | None = None
    state: str | None = None
    zip_code: str | None = None
    territory_id: str | None = None
    insurance: IntakeInsurancePayload
    physician: IntakePhysicianPayload | None = None
    order: dict[str, Any]


class EligibilityRequestPayload(BaseModel):
    order_id: str
    patient_id: str
    insurance_id: str | None = None
    payer_id: str


class PriorAuthRequestPayload(BaseModel):
    order_id: str
    payer_id: str
    hcpcs_codes: list[str] = []
    icd10_codes: list[str] = []
    requested_units: int = 1


async def _availity_token() -> str:
    if not settings.availity_client_id or not settings.availity_client_secret:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Availity credentials are not configured.",
        )
    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.post(
            settings.availity_token_url or "https://api.availity.com/availity/v1/token",
            data={
                "grant_type": "client_credentials",
                "client_id": settings.availity_client_id,
                "client_secret": settings.availity_client_secret,
            },
            headers={"Accept": "application/json"},
        )
    if response.status_code >= 400:
        logger.warning("Availity token request failed: status=%s", response.status_code)
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="Availity token request failed.",
        )
    token = response.json().get("access_token")
    if not token:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="Availity token response missing access token.",
        )
    return token


async def _run_eligibility_check(conn, payload: EligibilityRequestPayload) -> dict[str, Any]:
    token = await _availity_token()
    request_payload = {
        "payer": {"id": payload.payer_id},
        "patient": {"id": payload.patient_id},
        "serviceDates": [datetime.now(timezone.utc).date().isoformat()],
    }
    async with httpx.AsyncClient(timeout=45.0) as client:
        response = await client.post(
            f"{settings.availity_base_url.rstrip('/')}/availity/v1/eligibility-and-benefits-inquiries",
            json=request_payload,
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        )
    if response.status_code >= 400:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="Eligibility check failed with Availity.",
        )
    response_payload = response.json()
    is_eligible = bool(response_payload.get("eligible", True))
    status_value = "eligible" if is_eligible else "ineligible"

    check_id = str(uuid.uuid4())
    await exec_write(
        conn,
        """
        INSERT INTO eligibility_checks (id, order_id, patient_id, insurance_id, payer_id, status, is_eligible, coverage_details, raw_response)
        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9)
        """,
        check_id,
        payload.order_id,
        payload.patient_id,
        payload.insurance_id,
        payload.payer_id,
        status_value,
        is_eligible,
        json.dumps(response_payload),
        json.dumps(response_payload),
    )
    if payload.insurance_id:
        await exec_write(
            conn,
            """
            UPDATE patient_insurances
            SET eligibility_status = $1, eligibility_checked_at = NOW(), eligibility_response = $2
            WHERE id = $3
            """,
            status_value,
            json.dumps(response_payload),
            payload.insurance_id,
        )
    await exec_write(
        conn,
        """
        UPDATE orders
        SET status = CASE WHEN $1 = 'eligible' THEN status ELSE 'eligibility_failed' END,
            eligibility_status = $2,
            eligibility_summary = $3,
            updated_at = NOW()
        WHERE id = $4
        """,
        status_value,
        status_value,
        json.dumps(response_payload),
        payload.order_id,
    )
    return {
        "id": check_id,
        "status": status_value,
        "is_eligible": is_eligible,
        "response": response_payload,
    }


async def _run_prior_auth_check(conn, org_id: str, payload: PriorAuthRequestPayload) -> dict[str, Any]:
    token = await _availity_token()
    async with httpx.AsyncClient(timeout=45.0) as client:
        response = await client.post(
            f"{settings.availity_base_url.rstrip('/')}/availity/v1/prior-authorization-requests",
            json=payload.model_dump(),
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        )
    if response.status_code >= 400:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="Prior authorization request failed with Availity.",
        )
    response_payload: dict[str, Any] = response.json()

    auth_id = str(uuid.uuid4())
    auth_number = response_payload.get("auth_number") or response_payload.get("authorizationNumber")
    status_value = response_payload.get("status") or "submitted"
    await exec_write(
        conn,
        """
        INSERT INTO auth_requests (
            id, order_id, org_id, payer_id, auth_number, status, hcpcs_codes, icd10_codes, requested_units, raw_response, submitted_at
        )
        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,NOW())
        """,
        auth_id,
        payload.order_id,
        org_id,
        payload.payer_id,
        auth_number,
        status_value,
        payload.hcpcs_codes,
        payload.icd10_codes,
        payload.requested_units,
        json.dumps(response_payload),
    )
    await exec_write(
        conn,
        """
        UPDATE orders
        SET status = CASE WHEN $1 IN ('approved','not_required') THEN 'auth_approved' ELSE 'pending_auth' END,
            insurance_auth_number = COALESCE($2, insurance_auth_number),
            updated_at = NOW()
        WHERE id = $3
        """,
        status_value,
        auth_number,
        payload.order_id,
    )
    return {"id": auth_id, "status": status_value, "auth_number": auth_number, "response": response_payload}


async def _auto_process_order(conn, org_id: str, order_id: str, patient_id: str, insurance_id: str | None, payer_id: str, hcpcs_codes: list[str], icd10_codes: list[str]):
    try:
        eligibility_result = await _run_eligibility_check(
            conn,
            EligibilityRequestPayload(order_id=order_id, patient_id=patient_id, insurance_id=insurance_id, payer_id=payer_id),
        )
        await _record_workflow_event(conn, org_id, "eligibility.auto_checked", {"order_id": order_id, "status": eligibility_result["status"]}, order_id=order_id)
    except HTTPException as exc:
        await exec_write(
            conn,
            """
            UPDATE orders
            SET status = 'eligibility_failed',
                eligibility_status = 'error',
                updated_at = NOW()
            WHERE id = $1
            """,
            order_id,
        )
        await _record_workflow_event(
            conn,
            org_id,
            "eligibility.auto_check_failed",
            {"order_id": order_id, "detail": str(exc.detail)},
            order_id=order_id,
        )
        return

    requirement = await fetch_one(
        conn,
        """
        SELECT COUNT(*) AS count
        FROM payer_auth_requirements
        WHERE payer_id = $1 AND hcpcs_code = ANY($2) AND requires_auth = true
        """,
        payer_id,
        hcpcs_codes or [""],
    )
    if int((requirement or {}).get("count") or 0) > 0:
        try:
            await _run_prior_auth_check(
                conn,
                org_id,
                PriorAuthRequestPayload(
                    order_id=order_id,
                    payer_id=payer_id,
                    hcpcs_codes=hcpcs_codes,
                    icd10_codes=icd10_codes,
                    requested_units=1,
                ),
            )
            await _record_workflow_event(conn, org_id, "prior_auth.auto_checked", {"order_id": order_id, "payer_id": payer_id}, order_id=order_id)
        except HTTPException as exc:
            await exec_write(
                conn,
                """
                UPDATE orders
                SET status = 'pending_auth',
                    updated_at = NOW()
                WHERE id = $1
                """,
                order_id,
            )
            await _record_workflow_event(
                conn,
                org_id,
                "prior_auth.auto_check_failed",
                {"order_id": order_id, "detail": str(exc.detail)},
                order_id=order_id,
            )


@app.post("/api/v1/intake/patient", status_code=201)
async def v1_patient_intake(payload: IntakePatientPayload, background_tasks: BackgroundTasks, request: Request):
    db = request.app.state.db_pool
    idempotency_key = _request_idempotency_key(request)
    normalized_snapshot = payload.model_dump()
    payload_hash = _json_hash(normalized_snapshot)
    async with db.connection() as conn:
        await conn.execute("BEGIN")
        try:
            if idempotency_key:
                existing_order = await fetch_one(
                    conn,
                    """
                    SELECT id, patient_id
                    FROM orders
                    WHERE org_id = $1::uuid
                      AND intake_payload->>'_intake_idempotency_key' = $2
                    ORDER BY created_at DESC
                    LIMIT 1
                    """,
                    payload.org_id,
                    idempotency_key,
                )
                if existing_order:
                    await conn.execute("COMMIT")
                    return {
                        "patient_id": str(existing_order["patient_id"]),
                        "insurance_id": None,
                        "physician_id": None,
                        "order_id": str(existing_order["id"]),
                        "idempotent_replay": True,
                    }

            existing_patient = await fetch_one(
                conn,
                """
                SELECT id
                FROM patients
                WHERE org_id = $1
                  AND lower(first_name) = lower($2)
                  AND lower(last_name) = lower($3)
                  AND COALESCE(date_of_birth, dob) = $4
                LIMIT 1
                """,
                payload.org_id,
                payload.first_name,
                payload.last_name,
                payload.date_of_birth,
            )
            patient_id = str(existing_patient["id"]) if existing_patient else str(uuid.uuid4())
            if not existing_patient:
                await exec_write(
                    conn,
                    """
                    INSERT INTO patients (
                        id, org_id, first_name, last_name, date_of_birth, dob, gender, phone, email,
                        address_line1, address_line2, city, state, zip_code, territory_id, address
                    )
                    VALUES ($1,$2,$3,$4,$5,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15)
                    """,
                    patient_id,
                    payload.org_id,
                    payload.first_name,
                    payload.last_name,
                    payload.date_of_birth,
                    payload.gender,
                    payload.phone,
                    payload.email,
                    payload.address_line1,
                    payload.address_line2,
                    payload.city,
                    payload.state,
                    payload.zip_code,
                    payload.territory_id,
                    json.dumps(
                        {
                            "line1": payload.address_line1,
                            "line2": payload.address_line2,
                            "city": payload.city,
                            "state": payload.state,
                            "zip": payload.zip_code,
                        }
                    ),
                )

            insurance_id = str(uuid.uuid4())
            await exec_write(
                conn,
                """
                INSERT INTO patient_insurances (
                    id, patient_id, payer_name, payer_id, member_id, group_number,
                    subscriber_name, subscriber_dob, relationship, is_primary
                )
                VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10)
                """,
                insurance_id,
                patient_id,
                payload.insurance.payer_name,
                payload.insurance.payer_id,
                payload.insurance.member_id,
                payload.insurance.group_number,
                payload.insurance.subscriber_name,
                payload.insurance.subscriber_dob,
                payload.insurance.relationship,
                payload.insurance.is_primary,
            )

            physician_id = None
            if payload.physician and payload.physician.npi:
                existing_physician = await fetch_one(conn, "SELECT id FROM physicians WHERE npi = $1", payload.physician.npi)
                physician_id = str(existing_physician["id"]) if existing_physician else str(uuid.uuid4())
                if not existing_physician:
                    await exec_write(
                        conn,
                        """
                        INSERT INTO physicians (id, npi, first_name, last_name, specialty, phone, fax, facility_name, address)
                        VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9)
                        """,
                        physician_id,
                        payload.physician.npi,
                        payload.physician.first_name,
                        payload.physician.last_name,
                        payload.physician.specialty,
                        payload.physician.phone,
                        payload.physician.fax,
                        payload.physician.facility_name,
                        json.dumps({}),
                    )

            order_id = str(uuid.uuid4())
            order_number = f"INTAKE-{datetime.now(timezone.utc).strftime('%Y%m%d')}-{str(order_id)[:6].upper()}"
            order_data = payload.order
            hcpcs_codes = [item.get("hcpcs_code") for item in order_data.get("line_items", []) if item.get("hcpcs_code")]
            icd10_codes = [item.get("icd10_code") for item in order_data.get("diagnoses", []) if item.get("icd10_code")]
            intake_payload = dict(order_data)
            intake_payload["_intake_received_at"] = datetime.now(timezone.utc).isoformat()
            intake_payload["_intake_idempotency_key"] = idempotency_key
            intake_payload["_normalized_payload_hash"] = payload_hash
            intake_payload["_processing_status"] = "received"
            intake_payload["_processing_failure_reason"] = None

            await exec_write(
                conn,
                """
                INSERT INTO orders (
                    id, org_id, order_number, patient_id, physician_id, referring_physician_npi,
                    status, product_category, vertical, source, priority, territory_id, place_of_service,
                    clinical_notes, clinical_data, total_billed, date_of_service, hcpcs_codes, intake_payload
                )
                VALUES ($1,$2,$3,$4,$5,$6,'intake',$7,$8,$9,$10,$11,$12,$13,$14,$15,$16,$17,$18)
                """,
                order_id,
                payload.org_id,
                order_number,
                patient_id,
                physician_id,
                payload.physician.npi if payload.physician else None,
                order_data.get("product_category"),
                order_data.get("vertical"),
                order_data.get("source") or "intake_form",
                order_data.get("priority") or "normal",
                order_data.get("territory_id"),
                order_data.get("place_of_service") or "12",
                order_data.get("clinical_notes"),
                json.dumps(order_data.get("clinical_data") or {}),
                order_data.get("total_billed"),
                order_data.get("date_of_service"),
                json.dumps(hcpcs_codes),
                json.dumps(intake_payload),
            )
            for diagnosis in order_data.get("diagnoses", []):
                await exec_write(
                    conn,
                    """
                    INSERT INTO order_diagnoses (id, order_id, icd10_code, description, is_primary, sequence)
                    VALUES ($1,$2,$3,$4,$5,$6)
                    """,
                    str(uuid.uuid4()),
                    order_id,
                    diagnosis.get("icd10_code"),
                    diagnosis.get("description"),
                    bool(diagnosis.get("is_primary")),
                    diagnosis.get("sequence") or 1,
                )
            for item in order_data.get("line_items", []):
                await exec_write(
                    conn,
                    """
                    INSERT INTO order_line_items (id, order_id, hcpcs_code, modifier, description, quantity, unit_price, billed_amount)
                    VALUES ($1,$2,$3,$4,$5,$6,$7,$8)
                    """,
                    str(uuid.uuid4()),
                    order_id,
                    item.get("hcpcs_code"),
                    item.get("modifier"),
                    item.get("description"),
                    item.get("quantity") or 1,
                    item.get("unit_price"),
                    item.get("billed_amount"),
                )
            await _record_workflow_event(
                conn,
                payload.org_id,
                "intake.created",
                {
                    "order_id": order_id,
                    "patient_id": patient_id,
                    "source": order_data.get("source") or "intake_form",
                    "received_at": intake_payload["_intake_received_at"],
                    "idempotency_key": idempotency_key,
                    "payload_hash": payload_hash,
                },
                order_id=order_id,
            )
            await _auto_process_order(
                conn,
                payload.org_id,
                order_id,
                patient_id,
                insurance_id,
                payload.insurance.payer_id or "unknown",
                hcpcs_codes,
                icd10_codes,
            )
            await conn.execute("COMMIT")
        except Exception as exc:
            await conn.execute("ROLLBACK")
            logger.exception("Patient intake failed.")
            raise HTTPException(
                status_code=500,
                detail=f"Intake processing failed: {str(exc)[:200]}",
            )
    return {"patient_id": patient_id, "insurance_id": insurance_id, "physician_id": physician_id, "order_id": order_id, "idempotent_replay": False}


@app.post("/api/v1/intake/batch")
async def v1_intake_batch(file: UploadFile = File(...)):
    content = await file.read()
    filename = file.filename or "batch.csv"
    rows: list[dict[str, Any]] = []
    if filename.lower().endswith(".csv"):
        reader = csv.DictReader(io.StringIO(content.decode("utf-8", errors="replace")))
        rows = list(reader)
    elif filename.lower().endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        values = list(ws.iter_rows(values_only=True))
        headers = [str(value or "").strip() for value in (values[0] if values else [])]
        for record in values[1:]:
            rows.append({headers[i]: record[i] if i < len(record) else "" for i in range(len(headers))})
    else:
        raise HTTPException(status_code=400, detail="Accepted formats: csv, xlsx")

    processed = 0
    failed = 0
    errors: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        if not row:
            continue
        if not row.get("first_name") and not row.get("patient_name"):
            failed += 1
            errors.append({"row": index, "error": "Missing patient name"})
            continue
        processed += 1
    return {"processed": processed, "failed": failed, "errors": errors}


async def require_intake_internal_key(request: Request) -> bool:
    expected = (settings.internal_api_key or "").strip()
    if not expected:
        raise HTTPException(status_code=503, detail="INTERNAL_API_KEY is not configured")
    if request.headers.get("X-Internal-API-Key", "").strip() != expected:
        raise HTTPException(status_code=401, detail="Invalid internal service key")
    return True


def _extract_patient_name(text: str) -> str:
    """Extract patient name from EMR/medical PDFs with multiple fallback strategies."""
    # Strategy 1: "LAST, FIRST DOB:" header (eClinicalWorks repeating header)
    m = re.search(r"^([A-Z][A-Za-z'\-]+,\s*[A-Z][A-Za-z'\-]+)\s+(?:DOB|Date of Birth)\b", text, re.M)
    if m:
        return m.group(1).strip()

    # Strategy 2: "Patient Medical Record" header followed by name on next line
    m = re.search(r"Patient\s+Medical\s+Record\s*\n\s*([A-Z][A-Za-z'\-]+,\s*[A-Z][A-Za-z'\- ]+)", text, re.I)
    if m:
        name = m.group(1).strip()
        name = re.split(r"\d", name)[0].strip().rstrip(",")
        if name:
            return name

    # Strategy 3: "LAST, FIRST  ##Y, M/F" pattern (age/sex after name)
    m = re.search(r"^([A-Z][A-Za-z'\-]+,\s*[A-Z][A-Za-z'\- ]+?)\s+\d{1,3}\s*(?:Y|yo)\b", text, re.M | re.I)
    if m:
        return m.group(1).strip()

    # Strategy 4: "Patient Name:" / "Member Name:" explicit label
    m = re.search(r"(?:Patient|Member)\s+Name[:\s]+([A-Z][A-Za-z,'\-\s]+?)(?:\s{2,}|\t|\n|$)", text, re.I)
    if m:
        return m.group(1).strip()

    # Strategy 5: "Name:" label
    m = re.search(r"\bName[:\s]+([A-Z][A-Za-z,'\-\s]+?)(?:\s{2,}|\t|\n|$)", text, re.I)
    if m:
        return m.group(1).strip()

    return ""


def _extract_dob(text: str) -> str:
    """Extract date of birth, supporting numeric and month-name formats."""
    # Numeric: DOB: 06/13/1957 or DOB: 1957-06-13
    m = re.search(r"(?:DOB|Date of Birth|Birth Date)[:\s]+(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})", text, re.I)
    if m:
        return m.group(1).strip()

    # Month name: DOB: Jun 13, 1957
    m = re.search(
        r"(?:DOB|Date of Birth|Birth Date)[:\s]+"
        r"((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{1,2},?\s+\d{4})",
        text,
        re.I,
    )
    if m:
        return m.group(1).strip()

    # ISO: DOB: 1957-06-13
    m = re.search(r"(?:DOB|Date of Birth|Birth Date)[:\s]+(\d{4}-\d{2}-\d{2})", text, re.I)
    if m:
        return m.group(1).strip()

    return ""


def _extract_insurance(text: str) -> tuple[str, str]:
    """Extract payer name and member/subscriber ID."""
    payer_name = ""
    member_id = ""

    # Payer from "Primary Insurance: Humana" (stop at tab, double-space, or newline)
    m = re.search(
        r"(?:Primary|Secondary)?\s*Insurance[:\s]+([A-Za-z][A-Za-z0-9 &.'\-]+?)(?:\s{2,}|\t|\n|$)",
        text, re.I,
    )
    if m:
        payer_name = m.group(1).strip()

    if not payer_name:
        m = re.search(r"(?:Payer|Plan|Carrier)[:\s]+([A-Za-z][A-Za-z0-9 &.'\-]+?)(?:\s{2,}|\t|\n|$)", text, re.I)
        if m:
            payer_name = m.group(1).strip()

    # "Insurance: Humana Payer ID: 61101" pattern (eClinicalWorks encounter footer)
    if not payer_name:
        m = re.search(r"Insurance:\s*([A-Za-z][A-Za-z0-9 &.'\-]+?)\s+Payer\s+ID:", text, re.I)
        if m:
            payer_name = m.group(1).strip()

    # Member/Subscriber ID patterns
    m = re.search(r"(?:Member\s*ID|Subscriber\s*(?:No|Number|ID)|Insurance\s*ID)[:\s]+([A-Z0-9][A-Z0-9\-]+)", text, re.I)
    if m:
        member_id = m.group(1).strip()

    if not member_id:
        m = re.search(r"(?:Policy\s*Number|Member\s*#|ID\s*Number)[:\s]+([A-Z0-9][A-Z0-9\-]+)", text, re.I)
        if m:
            member_id = m.group(1).strip()

    return payer_name, member_id


@app.post("/api/v1/intake/parse-document")
async def v1_parse_document(
    request: Request,
    file: UploadFile = File(...),
    _: bool = Depends(require_intake_internal_key),
):
    cid = getattr(request.state, "correlation_id", None)
    logger.info("intake_parse_document correlation_id=%s filename=%s", cid, file.filename)
    content = await file.read()
    text = _extract_pdf_text(content)

    patient_name = _extract_patient_name(text)
    first_name, last_name = _split_name(patient_name)

    dob_raw = _extract_dob(text)

    payer_name, member_id = _extract_insurance(text)

    # ICD-10 diagnosis codes
    diagnosis_codes = list(set(re.findall(r"\b[A-Z]\d{2}(?:\.\d{1,4})?\b", text)))

    # HCPCS codes (L-codes, K-codes, E-codes, A-codes)
    hcpcs_codes = list(set(re.findall(r"\b[LKEAHJ]\d{4}\b", text)))

    # Infer HCPCS from ICD-10 when no explicit HCPCS found
    inferred_source = None
    if not hcpcs_codes and diagnosis_codes:
        hcpcs_codes, inferred_source = _infer_hcpcs_from_diagnosis(diagnosis_codes)
        if inferred_source and inferred_source != "no-inference":
            logger.info("Inferred HCPCS %s from ICD-10 via %s", ",".join(hcpcs_codes), inferred_source)

    physician_npi = _extract_first(text, [r"\bNPI[:\s]+(\d{10})\b"])

    parse_confidence = max(
        0.0,
        min(
            1.0,
            (
                (1.0 if patient_name else 0.0)
                + (1.0 if dob_raw else 0.0)
                + (1.0 if (payer_name or member_id) else 0.0)
                + (0.5 if hcpcs_codes else 0.0)
            )
            / 3.5,
        ),
    )

    logger.info(
        "intake_parse_document result: patient=%s dob=%s payer=%s member_id=%s dx=%s hcpcs=%s confidence=%.2f",
        patient_name, dob_raw, payer_name, member_id,
        ",".join(diagnosis_codes[:5]), ",".join(hcpcs_codes[:5]), parse_confidence,
    )

    return {
        "patient_name": patient_name,
        "first_name": first_name,
        "last_name": last_name,
        "date_of_birth": _normalize_date(dob_raw) if dob_raw else "",
        "insurance_info": {
            "payer_name": payer_name,
            "member_id": member_id,
        },
        "diagnosis_codes": diagnosis_codes,
        "physician_npi": physician_npi,
        "hcpcs_codes": hcpcs_codes,
        "hcpcs_inferred_from": inferred_source if inferred_source and inferred_source != "no-inference" else None,
        "raw_text_preview": text[:1500],
        "confidence": parse_confidence,
    }


@app.post("/api/v1/intake/eob")
async def v1_ingest_eob(request: Request, file: UploadFile | None = File(None), edi_text: str | None = Form(None)):
    db = request.app.state.db_pool
    if file is None and not edi_text:
        raise HTTPException(status_code=400, detail="Provide file or edi_text")
    if file is not None:
        content = await file.read()
        parsed = eob_parser.parse_eob_pdf(content) if (file.filename or "").lower().endswith(".pdf") else eob_parser.parse_835_edi(content.decode("utf-8", errors="replace"))
        raw_source = file.filename or "upload"
    else:
        parsed = eob_parser.parse_835_edi(edi_text or "")
        raw_source = "edi_text"

    created_claims = []
    async with db.connection() as conn:
        for claim in parsed.get("claims", [])[:100]:
            claim_id = str(uuid.uuid4())
            claim_number = claim.get("claim_id") or claim.get("payer_claim_number") or f"CLM-{claim_id[:8]}"
            denial_category = None
            denial_reason = None
            carc_code = None
            if claim.get("adjustments"):
                first_adjustment = claim["adjustments"][0]
                carc_code = first_adjustment.get("carc")
                denial_category = first_adjustment.get("category")
                denial_reason = denial_category
            appeal_deadline = (datetime.now(timezone.utc).date() + timedelta(days=60)).isoformat() if denial_category else None
            await exec_write(
                conn,
                """
                INSERT INTO eob_claims (
                    id, claim_number, patient_name, payer_id, provider_npi, date_of_service, total_billed, total_paid,
                    claim_status, denial_reason, carc_code, denial_category, suggested_action, is_appealable, appeal_deadline, raw_source
                )
                VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16)
                ON CONFLICT (claim_number) DO NOTHING
                """,
                claim_id,
                claim_number,
                f"{claim.get('patient_first','')} {claim.get('patient_last','')}".strip() or None,
                None,
                claim.get("provider_npi"),
                _normalize_date(claim.get("service_date")),
                claim.get("billed_amount"),
                claim.get("paid_amount"),
                claim.get("claim_status"),
                denial_reason,
                carc_code,
                denial_category,
                "review_denial" if denial_category else "post_payment",
                bool(denial_category),
                appeal_deadline,
                raw_source,
            )
            if claim.get("service_lines"):
                for line in claim["service_lines"]:
                    await exec_write(
                        conn,
                        """
                        INSERT INTO eob_line_items (id, claim_id, hcpcs_code, billed_amount, paid_amount, service_date)
                        VALUES ($1,$2,$3,$4,$5,$6)
                        """,
                        str(uuid.uuid4()),
                        claim_id,
                        line.get("hcpcs"),
                        line.get("billed"),
                        line.get("paid"),
                        _normalize_date(claim.get("service_date")),
                    )
            await exec_write(
                conn,
                """
                INSERT INTO eob_worklist (id, claim_id, status, action_type, notes, due_date)
                VALUES ($1,$2,'open',$3,$4,$5)
                """,
                str(uuid.uuid4()),
                claim_id,
                "appeal" if denial_category else "review",
                denial_reason or "Review remittance",
                appeal_deadline,
            )
            created_claims.append(claim_number)
    return {"claims_created": len(created_claims), "claim_numbers": created_claims, "parsed": parsed}


@app.post("/api/v1/intake/ar-import")
async def v1_ar_import(file: UploadFile = File(...)):
    content = await file.read()
    filename = file.filename or "ar.csv"
    rows: list[dict[str, Any]] = []
    if filename.lower().endswith(".csv"):
        rows = list(csv.DictReader(io.StringIO(content.decode("utf-8", errors="replace"))))
    elif filename.lower().endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        values = list(ws.iter_rows(values_only=True))
        headers = [str(v or "").strip() for v in (values[0] if values else [])]
        for record in values[1:]:
            rows.append({headers[i]: record[i] if i < len(record) else "" for i in range(len(headers))})
    else:
        raise HTTPException(status_code=400, detail="Accepted formats: csv, xlsx")
    return {"processed": len(rows), "mapped_to": "payment_outcomes"}


@app.post("/api/v1/intake/eligibility-check")
async def v1_eligibility_check(payload: EligibilityRequestPayload, request: Request):
    db = request.app.state.db_pool
    async with db.connection() as conn:
        result = await _run_eligibility_check(conn, payload)
    return result


@app.post("/api/v1/intake/prior-auth-check")
async def v1_prior_auth_check(payload: PriorAuthRequestPayload, request: Request):
    db = request.app.state.db_pool
    async with db.connection() as conn:
        order = await fetch_one(conn, "SELECT org_id FROM orders WHERE id = $1", payload.order_id)
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        result = await _run_prior_auth_check(conn, str(order["org_id"]), payload)
    return result


def _build_stedi_claim_payload(order: dict[str, Any], patient: dict[str, Any], line_items: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "controlNumber": str(uuid.uuid4())[:12],
        "claimInformation": {
            "patientControlNumber": order.get("order_number") or str(order["id"]),
            "payerClaimControlNumber": order.get("insurance_auth_number"),
            "totalClaimChargeAmount": float(order.get("total_billed") or 0),
        },
        "subscriber": {
            "memberId": patient.get("insurance_id"),
            "name": {"first": patient.get("first_name"), "last": patient.get("last_name")},
        },
        "provider": {"npi": order.get("referring_physician_npi")},
        "serviceLines": [
            {
                "procedureCode": item.get("hcpcs_code"),
                "chargeAmount": float(item.get("billed_amount") or 0),
                "quantity": int(item.get("quantity") or 1),
            }
            for item in line_items
        ],
    }


@app.post("/api/v1/intake/submit-claim/{order_id}")
async def v1_submit_claim(order_id: str, request: Request):
    db = request.app.state.db_pool
    idempotency_key = _request_idempotency_key(request)
    async with db.connection() as conn:
        if idempotency_key:
            existing = await fetch_one(
                conn,
                """
                SELECT id, acknowledgment_payload
                FROM claim_submissions
                WHERE order_id = $1
                  AND acknowledgment_payload->>'idempotency_key' = $2
                ORDER BY created_at DESC
                LIMIT 1
                """,
                order_id,
                idempotency_key,
            )
            if existing:
                return {"claim_submission_id": str(existing["id"]), "idempotent_replay": True}
        order = await fetch_one(conn, "SELECT * FROM orders WHERE id = $1", order_id)
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        patient = await fetch_one(conn, "SELECT * FROM patients WHERE id = $1", order["patient_id"])
        lines = [dict(row) for row in await fetch_all(conn, "SELECT * FROM order_line_items WHERE order_id = $1", order_id)]
        submission_payload = _build_stedi_claim_payload(dict(order), dict(patient), lines)
        stedi_api_key = os.getenv("STEDI_API_KEY", "").strip()
        response_payload: dict[str, Any]
        status_value = "submitted"
        if not stedi_api_key:
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="Stedi API key is not configured.",
            )
        # Stedi: raw API key in Authorization (not Bearer). Path matches EDI Stedi client professional claims v3.
        auth_header = stedi_api_key[7:].strip() if stedi_api_key.lower().startswith("bearer ") else stedi_api_key
        idem = idempotency_key or str(uuid.uuid4())
        async with httpx.AsyncClient(timeout=45.0) as client:
            response = await client.post(
                "https://healthcare.us.stedi.com/2024-04-01/change/medicalnetwork/professionalclaims/v3/submission",
                json=submission_payload,
                headers={
                    "Authorization": auth_header,
                    "Content-Type": "application/json",
                    "Idempotency-Key": idem,
                },
            )
        if response.status_code >= 400:
            raise HTTPException(
                status_code=status.HTTP_502_BAD_GATEWAY,
                detail="Claim submission failed with Stedi.",
            )
        response_payload = response.json()
        submission_id = str(uuid.uuid4())
        ack_payload = dict(response_payload)
        ack_payload["idempotency_key"] = idempotency_key
        claim_ref = response_payload.get("claimReference") or {}
        stedi_tx_id = (
            claim_ref.get("correlationId")
            or response_payload.get("transactionId")
            or response_payload.get("correlationId")
            or ""
        )
        stedi_status = str(response_payload.get("status") or "")
        if stedi_status == "SUCCESS":
            status_value = "accepted"
        await exec_write(
            conn,
            """
            INSERT INTO claim_submissions (
                id, order_id, org_id, payer_id, stedi_transaction_id, submission_payload, acknowledgment_payload, status, submitted_at
            )
            VALUES ($1,$2,$3,$4,$5,$6,$7,$8,NOW())
            """,
            submission_id,
            order_id,
            order["org_id"],
            order.get("payer_id"),
            stedi_tx_id,
            json.dumps(submission_payload),
            json.dumps(ack_payload),
            status_value,
        )
        if status_value == "submitted":
            await exec_write(conn, "UPDATE orders SET status = 'submitted', submitted_at = NOW(), updated_at = NOW() WHERE id = $1", order_id)
        await _record_workflow_event(conn, str(order["org_id"]), "claim.submitted", {"order_id": order_id}, order_id=order_id)
    return {"claim_submission_id": submission_id, "response": response_payload, "idempotent_replay": False}


@app.get("/api/v1/intake/claim-status/{order_id}")
async def v1_claim_status(order_id: str, request: Request):
    db = request.app.state.db_pool
    async with db.connection() as conn:
        row = await fetch_one(
            conn,
            """
            SELECT id, stedi_transaction_id, status, acknowledgment_payload, submitted_at, acknowledged_at
            FROM claim_submissions
            WHERE order_id = $1
            ORDER BY created_at DESC
            LIMIT 1
            """,
            order_id,
        )
    if not row:
        raise HTTPException(status_code=404, detail="Claim submission not found")
    return _serialize(dict(row))


@app.post("/api/v1/intake/process-acknowledgment")
async def v1_process_acknowledgment(payload: dict[str, Any], request: Request):
    submission_id = payload.get("claim_submission_id")
    if not submission_id:
        raise HTTPException(status_code=400, detail="Missing claim_submission_id")
    db = request.app.state.db_pool
    async with db.connection() as conn:
        submission = await fetch_one(
            conn,
            "SELECT id, order_id, org_id, status FROM claim_submissions WHERE id = $1",
            submission_id,
        )
        if not submission:
            raise HTTPException(status_code=404, detail="Claim submission not found")
        updated = await exec_write(
            conn,
            """
            UPDATE claim_submissions
            SET acknowledgment_payload = $1,
                status = COALESCE($2, status),
                acknowledged_at = NOW()
            WHERE id = $3
            """,
            json.dumps(payload),
            payload.get("status"),
            submission_id,
        )
        next_status = str(payload.get("status") or "").lower()
        if next_status in {"accepted", "paid"}:
            await exec_write(
                conn,
                "UPDATE orders SET status = 'pending_payment', updated_at = NOW() WHERE id = $1",
                str(submission["order_id"]),
            )
        elif next_status in {"denied", "rejected", "failed"}:
            await exec_write(
                conn,
                "UPDATE orders SET status = 'denied', updated_at = NOW() WHERE id = $1",
                str(submission["order_id"]),
            )
        await _record_workflow_event(
            conn,
            str(submission["org_id"]),
            "claim.acknowledgment_processed",
            {
                "claim_submission_id": submission_id,
                "order_id": str(submission["order_id"]),
                "previous_submission_status": str(submission["status"]),
                "ack_status": payload.get("status"),
                "processed_at": datetime.now(timezone.utc).isoformat(),
            },
            order_id=str(submission["order_id"]),
        )
    if not updated:
        raise HTTPException(status_code=404, detail="Claim submission not found")
    return {"status": "processed", "claim_submission_id": submission_id}
